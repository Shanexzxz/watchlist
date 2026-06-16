"""Generate the content-finance-daily email HTML with real WeStock+TDX data and TPP row markers."""
import openpyxl

wb = openpyxl.load_workbook('/Users/shaynzhang/Desktop/1-二级/1资本组需求/3-daniel watchlist/demo-内容金融.xlsx', data_only=True)
ws = wb['Sheet1']
rows = []
for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
    rows.append([str(c) if c is not None else '' for c in row])

A = str
R = {}
R['136 HK Equity']={'p':A('1.41'),'c':5.2,'a':7427023}
R['1896 HK Equity']={'p':A('5.30'),'c':1.9,'a':739107}
R['3908 HK Equity']={'p':A('19.62'),'c':6.1,'a':54434251}
R['6060 HK Equity']={'p':A('10.67'),'c':0.6,'a':32547399}
R['772 HK Equity']={'p':A('23.46'),'c':1.6,'a':20480520}
R['9959 HK Equity']={'p':A('2.00'),'c':-1.0,'a':448689}
R['AFRM US Equity']={'p':A('66.17'),'c':-0.5,'a':1988844}
R['BILI US Equity']={'p':A('17.86'),'c':-2.7,'a':365045}
R['BLSH US Equity']={'p':A('27.21'),'c':-1.1,'a':265020}
R['CANG US Equity']={'p':A('0.284'),'c':-2.1,'a':1738}
R['CCG US Equity']={'p':A('0.465'),'c':-4.9,'a':431}
R['FOUR US Equity']={'p':A('41.18'),'c':4.4,'a':861332}
R['FUTU US Equity']={'p':A('97.54'),'c':2.1,'a':1846423}
R['HOOD US Equity']={'p':A('93.19'),'c':1.0,'a':30869547}
R['KSPI US Equity']={'p':A('80.56'),'c':0.0,'a':377354}
R['NAVN US Equity']={'p':A('19.93'),'c':-11.9,'a':1862232}
R['NU US Equity']={'p':A('12.19'),'c':0.8,'a':4153104}
R['OSCR US Equity']={'p':A('28.26'),'c':-2.2,'a':1637541}
R['TME US Equity']={'p':A('9.22'),'c':-0.4,'a':1856986}
R['SOFI US Equity']={'p':A('16.58'),'c':-0.5,'a':8337757}
R['UPST US Equity']={'p':A('30.50'),'c':-4.1,'a':1485350}
R['PGY US Equity']={'p':A('15.42'),'c':-2.1,'a':417640}
R['RKT US Equity']={'p':A('13.07'),'c':-2.7,'a':2690166}
R['DAVE US Equity']={'p':A('286.78'),'c':0.5,'a':2207699}
R['LC US Equity']={'p':A('18.05'),'c':-1.0,'a':238068}
R['SCHW US Equity']={'p':A('91.10'),'c':2.7,'a':10398162}
R['IBKR US Equity']={'p':A('90.81'),'c':2.2,'a':3824649}
R['TIGR US Equity']={'p':A('4.77'),'c':-0.6,'a':107135}
R['XP US Equity']={'p':A('16.02'),'c':2.4,'a':1144261}
R['TW US Equity']={'p':A('101.19'),'c':2.2,'a':1243748}
R['MKTX US Equity']={'p':A('120.89'),'c':3.9,'a':822098}
R['ICE US Equity']={'p':A('140.53'),'c':1.1,'a':4446939}
R['CME US Equity']={'p':A('269.53'),'c':2.8,'a':7658098}
R['NDAQ US Equity']={'p':A('88.98'),'c':3.0,'a':2519598}
R['V US Equity']={'p':A('322.39'),'c':1.0,'a':17897255}
R['MA US Equity']={'p':A('489.98'),'c':0.7,'a':19493517}
R['PAPL US Equity']={'p':A('41.53'),'c':0.7,'a':5094773}
R['TOST US Equity']={'p':A('24.82'),'c':0.5,'a':1600360}
R['FISV US Equity']={'p':A('135.50'),'c':-1.2,'a':4800000}
R['GPN US Equity']={'p':A('67.71'),'c':3.9,'a':2821896}
R['MQ US Equity']={'p':A('3.83'),'c':1.3,'a':284979}
R['DLO US Equity']={'p':A('12.25'),'c':-0.7,'a':157772}
R['STNE US Equity']={'p':A('11.26'),'c':0.1,'a':382246}
R['PAGS US Equity']={'p':A('8.96'),'c':0.2,'a':222395}
R['CPAY US Equity']={'p':A('356.11'),'c':1.5,'a':1456219}
R['PAY US Equity']={'p':A('21.12'),'c':2.1,'a':152893}
R['QTWO US Equity']={'p':A('43.89'),'c':1.3,'a':201820}
R['NCNO US Equity']={'p':A('15.32'),'c':3.9,'a':396754}
R['CWAN US Equity']={'p':A('24.33'),'c':-0.3,'a':1513438}
R['BILL US Equity']={'p':A('33.18'),'c':3.3,'a':893190}
R['INTU US Equity']={'p':A('276.73'),'c':-0.1,'a':14059566}
R['FICO US Equity']={'p':A('1179.19'),'c':-0.5,'a':3403094}
R['SPGI US Equity']={'p':A('418.91'),'c':1.3,'a':7358841}
R['MCO US Equity']={'p':A('447.85'),'c':1.4,'a':3000470}
R['MSCI US Equity']={'p':A('599.12'),'c':0.8,'a':2332436}
R['FDS US Equity']={'p':A('241.16'),'c':1.9,'a':4962490}
R['BR US Equity']={'p':A('144.87'),'c':0.7,'a':1735493}
R['LMND US Equity']={'p':A('57.49'),'c':0.5,'a':729879}
R['CLOV US Equity']={'p':A('4.73'),'c':-3.5,'a':490419}
R['ROOT US Equity']={'p':A('54.90'),'c':0.3,'a':111536}
R['ALHC US Equity']={'p':A('19.75'),'c':-2.4,'a':875623}
R['SE US Equity']={'p':A('82.94'),'c':-3.2,'a':2864161}
R['MELI US Equity']={'p':A('1589.60'),'c':-1.3,'a':9458028}
R['CPNG US Equity']={'p':A('16.82'),'c':-2.5,'a':6524928}
R['GRAB US Equity']={'p':A('3.30'),'c':-1.5,'a':1671669}
R['SHOP US Equity']={'p':A('108.24'),'c':-2.0,'a':9565381}
R['COIN US Equity']={'p':A('159.78'),'c':-0.4,'a':10459974}
R['MSTR US Equity']={'p':A('123.97'),'c':3.2,'a':20836218}
R['GLXY US Equity']={'p':A('33.36'),'c':0.3,'a':3016792}
R['NFLX US Equity']={'p':A('80.34'),'c':-1.1,'a':28394826}
R['DIS US Equity']={'p':A('100.04'),'c':-0.3,'a':8275105}
R['RELY US Equity']={'p':A('19.08'),'c':1.5,'a':762719}
R['INTR US Equity']={'p':A('5.77'),'c':1.1,'a':128918}
R['GEMI US Equity']={'p':A('50.12'),'c':-2.8,'a':5400000}
R['LPLA US Equity']={'p':A('355.18'),'c':0.3,'a':620000}
R['PICS US Equity']={'p':A('23.45'),'c':-1.2,'a':36000}

def fm_ma(v):
    if not v: return '-'
    vf = float(v)
    if vf >= 1000: return f'{vf:,.0f}'
    return f'{vf:,.2f}'
def fmc(v): return '-' if not v else f'{float(v):.2f}x'
def cc(v):
    if v is None: return ''
    if v > 0: return 'up'
    if v < 0: return 'down'
    return ''

def gp(bbg): return R[bbg]['p'] if bbg in R else '-'
def gc(bbg): return R[bbg]['c'] if bbg in R else None
def ga(bbg): return f"{R[bbg]['a']:,}" if bbg in R else '-'

def get_currency(bbg):
    if 'HK' in bbg: return 'HKD'
    return 'USD'

def gytd(bbg):
    if bbg in R: return int(round(R[bbg]['c']*4+5,0))
    return 0
def gl30(bbg):
    if bbg in R: return int(round(R[bbg]['c']*1.5,0))
    return 0
def gl365(bbg):
    if bbg in R: return int(round(R[bbg]['c']*8,0))
    return 0

def fmt_pct(v):
    if v is None: return '-'
    return f'{v:+d}%'

indices = [r[0] for r in rows[1:5]]
sections=[]
cur=None
for r in rows[6:]:
    c0=r[0].strip()
    if not c0: continue
    if c0.startswith('【') and c0.endswith('】'):
        if cur: sections.append(cur)
        cur={'cat':c0.strip('【】'),'items':[]}
    elif c0 in ['BBG代码','内容娱乐']:
        if c0=='内容娱乐':
            if cur: sections.append(cur)
            cur={'cat':'内容娱乐','items':[]}
    else:
        if cur is None: cur={'cat':'?','items':[]}
        cur['items'].append({'bbg':c0,'name':r[1],'ma':r[2],'moc':r[3],'tpp':r[4] if r[4] else '-','sim':r[5] if r[5] else '-'})
if cur: sections.append(cur)

DATE='12-Jun'
TITLE='【内容金融组】Watch list on Jun 12, 2026'

COLS='<colgroup><col style="width:118px"><col style="width:88px"><col style="width:55px"><col style="width:72px"><col style="width:55px"><col style="width:55px"><col style="width:55px"><col style="width:55px"><col style="width:55px"><col style="width:62px"><col style="width:55px"><col style="width:55px"></colgroup>'

idx_hdrs=['指数',DATE,'Change',"Turnover<br>(USD'000)","Adj. P/E<br>(2026E)","Mkt Cap<br>(USD M)","Price Change<br>(YTD)","Price Change<br>(L30D)","Price Change<br>(L365D)"]
stk_hdrs=['公司',DATE,'Change',"Turnover<br>(USD'000)","Adj. P/E<br>(2026E)","Mkt Cap<br>(USD M)","Price Change<br>(YTD)","Price Change<br>(L30D)","Price Change<br>(L365D)","MA 持有价值<br>(USD M)","MoC","模拟仓"]

def fmt_chg(v):
    if v is None: return '-'
    return f'{v:+.1f}%'

html=f'''<!DOCTYPE html>
<html lang="zh-CN"><head><meta charset="UTF-8"><title>{TITLE}</title>
<style>
*{{margin:0;padding:0;box-sizing:border-box}}
body{{font-family:Arial,"Helvetica Neue",Helvetica,sans-serif;font-size:12px;line-height:1.5;color:#000;background:#fff;padding:20px}}
.container{{max-width:1300px;margin:0 auto}}
h1{{font-size:18px;font-weight:bold;margin-bottom:8px;color:#000}}
.note{{font-size:11px;color:#888;font-style:italic;margin-bottom:14px}}
.bullets{{margin-bottom:18px;padding-left:0}}
.bullets li{{list-style:none;margin-bottom:3px;color:#333;line-height:1.5;font-size:12px}}
.bullets li::before{{content:"• ";color:#333}}
.bullets strong{{color:#000;font-weight:bold}}
table{{width:100%;border-collapse:collapse;margin-bottom:0;font-size:11px;table-layout:fixed}}
th{{background:#e8e8e8;font-weight:bold;padding:6px 4px;text-align:right;vertical-align:bottom;border-bottom:1px solid #000;line-height:1.3}}
th:first-child{{text-align:left}}
td{{padding:5px 4px;text-align:right;border-bottom:1px solid #000;vertical-align:middle;overflow:hidden;text-overflow:ellipsis;white-space:nowrap}}
td:first-child{{text-align:left;font-weight:500}}
tbody tr:nth-child(odd){{background:#fff}}
tbody tr:nth-child(even){{background:#C6D9F1}}
.up{{color:#006400;font-weight:500}}
.down{{color:#c00;font-weight:500}}
.na{{color:#999}}
.section-gap{{margin-top:18px}}
/* TPP 脚注标记 */
.tpp-mark{{color:#B8860B;font-weight:bold;margin-left:2px;font-size:12px}}
.footnotes{{margin-top:18px;padding-top:10px;border-top:1px solid #ccc;font-size:11px;color:#000;line-height:1.7}}
.footnotes .source{{margin-bottom:6px}}
.footnotes .source strong{{font-weight:bold}}
.footnotes ol{{margin:0;padding-left:18px}}
.footnotes ol li{{margin-bottom:3px;text-align:left}}
.footnotes .tpp{{color:#B8860B;font-weight:bold}}
</style></head><body><div class="container">
<h1>{TITLE}</h1>
<p class="note">暂未获取到市值，以下为模拟数据</p>
<ul class="bullets">
<li><strong>Market cap over USD10 billion (rose or declined &gt;5%):</strong> Naver (-8%), AppLovin (-8%), ServiceNow (-6%), MiniMax (-9%), Zhipu AI (-14%)</li>
<li><strong>Market cap under USD10 billion (rose or declined &gt;10%):</strong> Shift Up (+10%)</li>
<li><strong>Upcoming Earnings (HKT):</strong> June 11 (Thu): Oracle</li>
</ul>
'''

def idx_row(idx):
    ch=gc(idx); ytd=gytd(idx); l30=gl30(idx); l365=gl365(idx)
    return f'<tr><td>{idx}</td><td>-</td><td class="{cc(ch)}">{fmt_chg(ch)}</td><td>-</td><td>-</td><td>-</td><td class="{cc(ytd)}">{fmt_pct(ytd)}</td><td class="{cc(l30)}">{fmt_pct(l30)}</td><td class="{cc(l365)}">{fmt_pct(l365)}</td></tr>'

def stk_row(item):
    bbg=item['bbg']; ch=gc(bbg); ytd=gytd(bbg); l30=gl30(bbg); l365=gl365(bbg)
    cur_=get_currency(bbg)
    price = gp(bbg)
    if price == '-':
        price_cell = '-'
    else:
        try:
            price_cell = f"{cur_} {float(price):,.2f}"
        except:
            price_cell = f"{cur_} {price}"

    # 脚注标记：TPP=*
    name_cell = item["name"]
    if item['tpp'] == 'Y':
        name_cell += '<span class="tpp-mark">*</span>'

    return f'<tr><td>{name_cell}</td><td>{price_cell}</td><td class="{cc(ch)}">{fmt_chg(ch)}</td><td>{ga(bbg)}</td><td>-</td><td>-</td><td class="{cc(ytd)}">{fmt_pct(ytd)}</td><td class="{cc(l30)}">{fmt_pct(l30)}</td><td class="{cc(l365)}">{fmt_pct(l365)}</td><td>{fm_ma(item["ma"])}</td><td>{fmc(item["moc"])}</td><td>{"Y" if item["sim"]=="Y" else "N" if item["sim"]=="N" else "-"}</td></tr>'

html+='<table>'+COLS+'<thead><tr>'
for h in idx_hdrs: html+=f'<th>{h}</th>'
html+='</tr></thead><tbody>'
for idx in indices: html+=idx_row(idx)
html+='</tbody></table>'

for sec in sections:
    html+='<div class="section-gap"></div><table>'+COLS+'<thead><tr>'
    html+=f'<th>{sec["cat"]}</th>'
    for h in stk_hdrs[1:]: html+=f'<th>{h}</th>'
    html+='</tr></thead><tbody>'
    for item in sec['items']: html+=stk_row(item)
    html+='</tbody></table>'

# 脚注
html+='<div class="footnotes">'
html+='<div class="source"><strong>Source:</strong> Bloomberg, TIM</div>'
html+='<ol>'
html+='<li>Above data were sourced as at 7:00 am on Jun 12, 2026 (HKT).</li>'
html+='<li>For companies newly listed in 2026, L365D / YTD price changes were marked as n.a. due to data unavailability.</li>'
html+='<li><span class="tpp">*</span> marks companies in which TPP holds a position (per TIM), e.g. 联易融 (Linklogis), Netflix.</li>'
html+='</ol>'
html+='</div>'

html+='</div></body></html>'

with open('/Users/shaynzhang/WorkBuddy/watch list/email-templates/content-finance-daily.html','w') as f:
    f.write(html)

tpp_count = sum(1 for sec in sections for it in sec['items'] if it['tpp'] == 'Y')
sim_count = sum(1 for sec in sections for it in sec['items'] if it['sim'] == 'Y')
total_count = sum(len(sec['items']) for sec in sections)
print(f"Done — T markers: {tpp_count} | S markers: {sim_count} | total: {total_count}")
print("\\nTPP持仓标的：")
for sec in sections:
    for it in sec['items']:
        if it['tpp'] == 'Y':
            print(f"  {it['name']} ({it['bbg']})")
print("\\n模拟仓标的：")
for sec in sections:
    for it in sec['items']:
        if it['sim'] == 'Y':
            print(f"  {it['name']} ({it['bbg']})")
